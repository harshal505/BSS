import openpyxl


def getRowCount(path, name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[name]
    return sheet.max_row


def getColumnCount(path, name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[name]
    return sheet.max_column


def readData(path, name, rowcount, column_count):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[name]
    return sheet.cell(row=rowcount, column=column_count).value


def writeData(path, name, rowcount, column_count, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[name]
    sheet.cell(row=rowcount, column=column_count).value = data
    workbook.save(path)
