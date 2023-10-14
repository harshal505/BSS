import openpyxl


class Read_test_data:
    @staticmethod
    def getData():
        fileName = r"H:\Harshal data\pythonProject\pythonProject2\Test_data\datafile.xlsx"
        book = openpyxl.load_workbook(fileName)
        sheet = book["User"]
        trows = sheet.max_row
        t_columns = sheet.max_column
        Dict = {}

        for r in range(2, 6):
            # rlist = []
            Dict.update(Dict)
            for c in range(1, 3):
                if c == 1:
                    Dict["Email"] = sheet.cell(row=r, column=c).value
                elif c == 2:
                    Dict["Password"] = sheet.cell(row=r, column=c).value
                print(Dict)
                Dict.update(Dict)

        return Dict


b = Read_test_data()
d = b.getData()
print(d)
