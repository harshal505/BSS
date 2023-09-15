import openpyxl

namefile = r"H:\Harshal data\pythonProject\pythonProject2\Test_data\datafile.xlsx"
workbook = openpyxl.load_workbook(namefile)
sheet = workbook['User']

# print(sheet.cell(row=1, column=1).value)
# print(sheet.cell(row=1, column=2).value)
r = sheet.max_row
c = sheet.max_column

for i in range(1, r+1):
    for itt in range(1, c+1):
        print(sheet.cell(row=i, column=itt).value)
