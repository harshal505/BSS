import openpyxl

file = r"H:\Harshal data\pythonProject\pythonProject2\Test_data\data1file.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["mer"]
row = sheet.max_row
col = sheet.max_column
a = {}
print(row, col)
for r in range(1, row + 1):
    for c in range(1, col + 1):
        print(sheet.cell(row=r, column=c).value)
        if c == 1:
            a["EMail"] = sheet.cell(row=r, column=c).value
        if c == 2:
            a["Password"] = sheet.cell(row=r, column=c).value
        print(a)
